import datetime
from pytz import timezone
from app.core.logger import logger
from app.database.postgres import get_db_cursor
from openpyxl.utils import (
    get_column_letter as openpyxl_get_column_letter,
    column_index_from_string,
)


def get_column_letter(col_idx):
    """Convert column index to Excel column letter (1-based)"""
    return openpyxl_get_column_letter(col_idx)


def format_date(date_str):
    """Convert UTC datetime string to IST date string"""
    if not date_str:
        return "N/A"
    try:
        # Check if date_str is already a datetime object
        if isinstance(date_str, str):
            # Parse the datetime string
            utc_dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        else:
            utc_dt = date_str
        # Convert to IST (UTC+5:30)
        ist = timezone("Asia/Kolkata")
        ist_dt = utc_dt.replace(tzinfo=timezone("UTC")).astimezone(ist)
        # Format as date string
        return ist_dt.strftime("%d/%m/%Y")
    except Exception as e:
        logger.error(f"Error formatting date: {e}")
        return date_str


def apply_border_to_range(worksheet, range_string, border):
    """
    Helper function to apply borders to all cells in a merged range
    """
    start_cell, end_cell = range_string.split(":")
    start_col = column_index_from_string("".join(filter(str.isalpha, start_cell)))
    start_row = int("".join(filter(str.isdigit, start_cell)))
    end_col = column_index_from_string("".join(filter(str.isalpha, end_cell)))
    end_row = int("".join(filter(str.isdigit, end_cell)))

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border


async def get_semester_id_from_class_id(class_id: str):
    """
    Query the database for the semester corresponding to a given class ID.

    Args:
        class_id (str): The ID of the class.

    Returns:
        The semester value if found, otherwise None.
    """
    with get_db_cursor() as (cursor, conn):
        cursor.execute('SELECT semester FROM "Class" WHERE id = %s', (class_id,))
        result = cursor.fetchone()
        if result and "semester" in result:
            return result["semester"]
        return None


async def get_semester_id_from_course_id(course_id: str):
    """
    Query the database for the semester corresponding to a given course ID.

    Args:
        course_id (str): The ID of the course.

    Returns:
        The semester value if found, otherwise None.
    """
    with get_db_cursor() as (cursor, conn):
        cursor.execute('SELECT "semesterId" FROM "Course" WHERE id = %s', (course_id,))
        result = cursor.fetchone()
        if result and result.get("semesterId") is not None:
            return result["semesterId"]
        else:
            # Get from class Id if semesterId is None when using courseId
            logger.warning(
                f"SemesterId is None for courseId {course_id}. Trying to get from classId."
            )
            cursor.execute('SELECT "classId" FROM "Course" WHERE id = %s', (course_id,))
            class_result = cursor.fetchone()
            if class_result and class_result.get("classId") is not None:
                return await get_semester_id_from_class_id(class_result["classId"])

        return None

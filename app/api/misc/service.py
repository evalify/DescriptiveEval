from app.database.postgres import get_db_cursor
import pandas as pd
from io import BytesIO
import os
from datetime import datetime

from app.core.logger import logger
from .utils import format_date, get_column_letter, apply_border_to_range
from typing import Dict


def populate_course_sheet(
    writer, course_id: str, report_data: dict, sheet_name: str = None
):
    """
    Helper function to populate a course report sheet in an Excel workbook.
    This reduces redundancy between generate_excel_report and generate_excel_class_report.

    Args:
        writer: pd.ExcelWriter - The Excel writer to use
        course_id: str - The course ID
        report_data: dict - The report data from fetch_course_report
        sheet_name: str - The sheet name to use (defaults to course_code if not provided)
    """
    # Create DataFrame for student data
    df = pd.DataFrame(report_data["students"])
    # Replace NaN values with "0"
    df = df.fillna(0)
    logger.debug("DataFrame created from student data.")

    # Get course and class details
    with get_db_cursor() as (cursor, conn):
        cursor.execute(
            'SELECT "name", "code" FROM "Course" WHERE "id" = %s', (course_id,)
        )
        course_result = cursor.fetchone()
        course_name = course_result["name"] if course_result else course_id
        course_code = course_result["code"] if course_result else ""

        cursor.execute(
            'SELECT "name" FROM "Class" WHERE "id" = %s', (report_data["class_id"],)
        )
        class_result = cursor.fetchone()
        class_name = class_result["name"] if class_result else report_data["class_id"]
    logger.debug("Course and class names fetched from database.")

    # Determine sheet name if not provided
    if not sheet_name:
        sheet_name = (
            course_code[:30] if course_code else course_id[:30]
        )  # Excel sheet name limit is 31 characters

    # ----- CALCULATE PERCENTAGES AND CREATE COMBINED DATAFRAME -----
    quiz_details = report_data["quiz_details"]

    # Create DataFrame for percentages (with the same quiz IDs as column names)
    percent_df = pd.DataFrame(index=df.index)

    # Create a dictionary to store the numerical percentage values for sorting
    percentage_values = {}

    # Calculate percentages for each quiz
    for quizId in quiz_details:
        total_score = quiz_details[quizId]["totalScore"]
        if total_score and total_score > 0:  # Avoid division by zero
            # Calculate regular percentages as numerical values (not strings)
            percent_df[f"{quizId} %"] = df[quizId].apply(
                lambda score: score / total_score
                if score != 0 or pd.notna(score)
                else 0
            )

            # Store numerical percentages for sorting (as decimal values, not percentages)
            percentage_values[quizId] = df[quizId].apply(
                lambda score: score / total_score
                if score != 0 or pd.notna(score)
                else 0
            )
        else:
            # Handle case where total_score is 0 or None
            percent_df[f"{quizId} %"] = "N/A"
            percentage_values[quizId] = 0

    # Create DataFrame for sorted percentages (with generic column names)
    sorted_percent_df = pd.DataFrame(index=df.index)

    # Calculate sorted percentages for each student
    for idx in range(len(df)):
        # Get percentage values for this student
        student_percentages = {
            quizId: percentage_values[quizId][idx] for quizId in quiz_details
        }

        # Sort percentages in descending order
        sorted_percents = sorted(
            student_percentages.items(), key=lambda x: x[1], reverse=True
        )

        # Add sorted percentages to DataFrame with generic column names (as numerical values)
        for i, (quizId, value) in enumerate(sorted_percents):
            # Ordinal names: "Best %", "2nd Best %", "3rd Best %", etc.
            if i == 0:
                column_name = "Best %"
            elif i == 1:
                column_name = "2nd Best %"
            elif i == 2:
                column_name = "3rd Best %"
            else:
                column_name = f"{i + 1}th Best %"

            if column_name not in sorted_percent_df.columns:
                sorted_percent_df[column_name] = 0.0
            sorted_percent_df.at[idx, column_name] = float(value)

    # Empty df for vertical spacing
    spacing_df = pd.DataFrame([""] * len(df), columns=["spacing"])
    spacing_offset = 1  # Offset for other columns that depend on spacing

    # Combine all DataFrames horizontally
    combined_df = pd.concat([df, spacing_df, percent_df, sorted_percent_df], axis=1)

    # Write the combined DataFrame to Excel
    df_start_row = 15
    combined_df.to_excel(
        writer, sheet_name=sheet_name, index=False, startrow=df_start_row
    )

    logger.debug(
        "Combined data (raw scores, percentages, sorted percentages) written to Excel sheet."
    )

    # Get the worksheet to apply formatting
    workbook = writer.book
    worksheet = workbook[sheet_name]

    # Add title and summary information
    from openpyxl.styles import Font, Border, PatternFill, Alignment, Side

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Set border to thin border for all dumped cells except spacing column
    n_rows, n_cols = combined_df.shape
    spacing_column_index = len(df.columns) + 1  # +1 because Excel columns are 1-indexed
    for r in range(df_start_row + 1, df_start_row + n_rows + 2):
        for c in range(1, n_cols + 1):
            if c != spacing_column_index:  # Skip applying border to spacing column
                worksheet.cell(row=r, column=c).border = thin_border

    # Title row: merged cells A1 to H1 for a prominent course title
    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = f"{course_code} - {course_name}"
    worksheet["A1"].font = Font(size=22, bold=True)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    logger.debug("Quiz Title row formatted.")

    # Prepare timeframe display text
    timeframe_text = "All quizzes"
    if report_data.get("start_date") and report_data.get("end_date"):
        timeframe_text = (
            f"From {report_data['start_date']} to {report_data['end_date']}"
        )
    elif report_data.get("start_date"):
        timeframe_text = f"From {report_data['start_date']} onwards"
    elif report_data.get("end_date"):
        timeframe_text = f"Until {report_data['end_date']}"

    # Summary section starting at row 2 with clear labels and values
    summary_data = [
        ("Course Code", course_code),
        ("Course Name", course_name),
        ("Class Name", class_name),
        ("Total Students", report_data["student_count"]),
        ("Total Quizzes", report_data["quiz_count"]),
        ("Timeframe", timeframe_text),
        ("Generated On", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("CourseId", course_id),
        ("ClassId", report_data["class_id"]),
    ]
    for i, (label, value) in enumerate(summary_data, start=2):
        label_cell = worksheet[f"A{i}"]
        label_cell.value = label
        label_cell.fill = PatternFill(
            start_color="FABF8F", end_color="FABF8F", fill_type="solid"
        )
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="left")
        label_cell.border = thin_border

        worksheet[f"B{i}"] = value
        worksheet[f"B{i}"].fill = PatternFill(
            start_color="C4BD97", end_color="C4BD97", fill_type="solid"
        )
        worksheet[f"B{i}"].alignment = Alignment(horizontal="left")
        worksheet[f"B{i}"].border = thin_border

    logger.debug("Summary data added to Excel sheet.")

    # Student data headers (and data) will start from row 12, as defined in df.to_excel

    # Make the main headers bold (row 12)
    for cell in worksheet[df_start_row + 1]:  # Row 12 has the student headers
        if cell.value in ["Student Name", "Roll Number"]:
            current_value = cell.value

            worksheet.merge_cells(
                start_row=cell.row - 4,
                start_column=cell.col_idx,
                end_row=cell.row,
                end_column=cell.col_idx,
            )

            cell = worksheet.cell(
                row=cell.row - 4,
                column=cell.col_idx,
            )
            cell.value = current_value
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(
                start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
            )
            # Merge with 4 cells above
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )

            cell.border = thin_border

        if "Best" in str(cell.value):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )
    logger.debug("Student data headers formatted.")

    # Quiz details header section
    start_col = 3  # starting from column C
    quiz_count = len(quiz_details)

    # Add section headings for Marks and Percentages (Row 11)
    # Marks section heading (spans all raw score columns)
    marks_title_row = 12
    marks_start_col = get_column_letter(start_col)
    marks_end_col = get_column_letter(start_col + quiz_count - 1)
    marks_range = f"{marks_start_col}{marks_title_row}:{marks_end_col}{marks_title_row}"
    worksheet.merge_cells(marks_range)
    worksheet[f"{marks_start_col}{marks_title_row}"] = "Marks"
    worksheet[f"{marks_start_col}{marks_title_row}"].font = Font(size=14, bold=True)
    worksheet[f"{marks_start_col}{marks_title_row}"].alignment = Alignment(
        horizontal="center", vertical="center"
    )
    worksheet[f"{marks_start_col}{marks_title_row}"].fill = PatternFill(
        start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
    )

    # Apply border to all cells in the merged marks range
    apply_border_to_range(worksheet, marks_range, thin_border)

    # Percentages section heading (spans all percentage and sorted percentage columns)
    percent_start_col = get_column_letter(start_col + quiz_count + spacing_offset)
    percent_end_col = get_column_letter(
        start_col + quiz_count * 3 - 1 + spacing_offset
    )  # Includes both regular and sorted percentages
    percent_range = (
        f"{percent_start_col}{marks_title_row}:{percent_end_col}{marks_title_row}"
    )
    worksheet.merge_cells(percent_range)
    worksheet[f"{percent_start_col}{marks_title_row}"] = "Percentages"
    worksheet[f"{percent_start_col}{marks_title_row}"].font = Font(size=14, bold=True)
    worksheet[f"{percent_start_col}{marks_title_row}"].alignment = Alignment(
        horizontal="center", vertical="center"
    )
    worksheet[f"{percent_start_col}{marks_title_row}"].fill = PatternFill(
        start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
    )
    apply_border_to_range(
        worksheet, percent_range, thin_border
    )  # Apply border to all cells in the merged percentages range

    # Place Quiz Titles in row 8 with contrasting style and wrap text enabled
    quiz_info_table_start_row = 13
    for idx, quizId in enumerate(quiz_details, start=start_col):
        # Original quiz columns
        col_letter = get_column_letter(idx)
        cell = worksheet[f"{col_letter}{quiz_info_table_start_row}"]
        cell.value = quiz_details[quizId]["title"]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        # Enable text wrapping to handle long quiz titles
        cell.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )

    # Place Quiz Dates in row 9
    for idx, quizId in enumerate(quiz_details, start=start_col):
        # Original date columns
        col_letter = get_column_letter(idx)
        worksheet[f"{col_letter}{quiz_info_table_start_row + 1}"] = quiz_details[
            quizId
        ]["date"]
        worksheet[f"{col_letter}{quiz_info_table_start_row + 1}"].font = Font(bold=True)
        worksheet[f"{col_letter}{quiz_info_table_start_row + 1}"].fill = PatternFill(
            start_color="FFD966", end_color="FFD966", fill_type="solid"
        )

    # Place Quiz Total Scores in row 10
    for idx, quizId in enumerate(quiz_details, start=start_col):
        # Original total score columns
        col_letter = get_column_letter(idx)
        worksheet[f"{col_letter}{quiz_info_table_start_row + 2}"] = (
            f"Total: {quiz_details[quizId]['totalScore']}"
        )
        worksheet[f"{col_letter}{quiz_info_table_start_row + 2}"].font = Font(bold=True)
        worksheet[f"{col_letter}{quiz_info_table_start_row + 2}"].fill = PatternFill(
            start_color="FFD966", end_color="FFD966", fill_type="solid"
        )
        worksheet[f"{col_letter}{quiz_info_table_start_row + 2}"].alignment = Alignment(
            horizontal="left"
        )

    # Add header for the sorted percentages section
    rank_start_col = start_col + quiz_count * 2 + spacing_offset
    # Apply percentage formatting to percentage columns and sorted percentage columns
    # Find where percentage columns start
    percent_start_col_num = start_col + quiz_count + spacing_offset

    # Apply formatting to each row in the percentage columns
    for row in range(
        df_start_row + 2, df_start_row + 2 + len(df)
    ):  # Start from row 13 (data rows after header)
        # Format regular percentage columns
        for i in range(quiz_count):
            col_letter = get_column_letter(percent_start_col_num + i)
            cell = worksheet[f"{col_letter}{row}"]
            cell.number_format = "0.00%"  # Format as percentage with 2 decimal places

        # Format sorted percentage columns
        for i in range(quiz_count):
            col_letter = get_column_letter(rank_start_col + i)
            cell = worksheet[f"{col_letter}{row}"]
            cell.number_format = "0.00%"  # Format as percentage with 2 decimal places

    # Auto-adjust column width
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column if hasattr(col[0], "column") else col[0].column_letter
        column_letter = get_column_letter(column) if isinstance(column, int) else column
        for cell in col:
            if cell.value == "spacing":
                cell.value = ""
                cell.border = Border()
                max_length = 20
            elif cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_letter].width = adjusted_width
    logger.debug("Column widths auto-adjusted.")

    return course_code


async def generate_excel_class_report(
    class_id: str,
    save_to_file: bool = True,
    start_date: str = None,
    end_date: str = None,
) -> Dict[str, BytesIO | str]:
    """
    Generate a multi-sheet Excel report for all courses in a class.
    Each course report will be a separate sheet in the Excel file.

    Args:
        class_id: str - The class ID
        save_to_file: bool - Whether to save the report to a file
        start_date: str - Optional start date to filter quizzes (format: YYYY-MM-DD)
        end_date: str - Optional end date to filter quizzes (format: YYYY-MM-DD)
    """
    # Get all courseids for the class
    query = """
    SELECT c."id" AS "courseId"
    FROM "Course" c
    WHERE c."classId" = %s;
    """
    query2 = """SELECT cl."name" FROM "Class" cl WHERE cl."id" = %s;"""
    with get_db_cursor() as (cursor, conn):
        cursor.execute(query, (class_id,))
        rows = cursor.fetchall()
        logger.debug(f"Query executed, {len(rows)} rows fetched.")
        cursor.execute(query2, (class_id,))
        class_name = cursor.fetchone()["name"]
        logger.debug(f"Class name fetched: {class_name}")
    course_ids = [row["courseId"] for row in rows]
    logger.debug(f"Course IDs fetched for class {class_id}: {course_ids}")

    if not course_ids:
        logger.debug("No courses found for this class.")
        return {}

    outfile = BytesIO()
    # Create a single Excel writer for multiple sheets
    with pd.ExcelWriter(outfile, engine="openpyxl") as writer:
        for course_id in course_ids:
            report_data = await fetch_course_report(course_id, start_date, end_date)
            populate_course_sheet(writer, course_id, report_data)
            logger.debug(f"Report sheet generated for course ID: {course_id}")

    # Reset file pointer to beginning
    outfile.seek(0)

    # Save the Excel file
    if save_to_file:
        # Create directory if it doesn't exist
        directory = "data/reports"
        os.makedirs(directory, exist_ok=True)

        # Create filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Include timeframe info in filename if present
        timeframe_text = ""
        if start_date or end_date:
            if start_date and end_date:
                timeframe_text = f"_{start_date}_to_{end_date}"
            elif start_date:
                timeframe_text = f"_from_{start_date}"
            elif end_date:
                timeframe_text = f"_until_{end_date}"

        filename = (
            f"{directory}/class_{class_id}{timeframe_text}_report_{timestamp}.xlsx"
        )

        # Save to file
        with open(filename, "wb") as file:
            file.write(outfile.getvalue())
        logger.debug(f"Excel report saved to {filename}")
        return filename

    logger.debug(f"Excel report created in memory for class ID: {class_id}")
    return {"file": outfile, "class_name": class_name}


async def fetch_course_report(
    course_id: str, start_date: str = None, end_date: str = None
):
    logger.debug(
        f"Fetching course report for course ID: {course_id}, timeframe: {start_date} to {end_date}"
    )
    query = """
    SELECT 
        s."id" AS "studentId",
        u."name" AS "studentName",  
        u."rollNo" AS "studentRollNo",
        s."classId",
        cq."B" AS "quizId",
        q."title" AS "quizTitle",
        q."startTime" AS "quizStartTime",
        qr."score",
        qr."totalScore"
    FROM "Student" s
    JOIN "User" u ON s."id" = u."id" AND u."role" = 'STUDENT'
    JOIN "Course" c ON s."classId" = c."classId"
    JOIN "_CourseToQuiz" cq ON c."id" = cq."A"
    LEFT JOIN "Quiz" q ON cq."B" = q."id"
    LEFT JOIN "QuizResult" qr ON s."id" = qr."studentId" AND cq."B" = qr."quizId"
    WHERE c."id" = %s AND q."isEvaluated" = 'EVALUATED' AND qr."isEvaluated" = 'EVALUATED'
    """

    # Add date filtering if timeframe parameters are provided
    params = [course_id]
    if start_date:
        query += ' AND q."startTime" >= %s'
        params.append(start_date)
    if end_date:
        query += ' AND q."startTime" <= %s'
        params.append(end_date)

    query += ' ORDER BY s."id", cq."B";'

    with get_db_cursor() as (cursor, conn):
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
        logger.debug(f"Query executed, {len(rows)} rows fetched.")

    unique_quiz_ids = set(row["quizId"] for row in rows)
    logger.debug(f"Unique quiz IDs fetched: {unique_quiz_ids}")

    # Dictionary to store student data
    students_data = {}
    quiz_info = {}
    class_id = None

    # First, gather all quizzes and their details
    for row in rows:
        quiz_id = row["quizId"]
        if quiz_id not in quiz_info:
            quiz_info[quiz_id] = {
                "title": row["quizTitle"] or f"Quiz {quiz_id}",
                "totalScore": row["totalScore"],
                "startTime": row["quizStartTime"],
            }
            logger.debug(f"Quiz ID {quiz_id} added to quiz_info.")
        else:
            # Check if total score is consistent
            if (
                row["totalScore"] is not None
                and quiz_info[quiz_id]["totalScore"] is not None
            ):
                if row["totalScore"] != quiz_info[quiz_id]["totalScore"]:
                    raise ValueError(
                        f"Inconsistent total scores for quiz {quiz_id}, check {row['studentId']}"
                    )

        if class_id is None and row["classId"]:
            class_id = row["classId"]
            logger.debug(f"Class ID set to {class_id}.")

    # Now process student data
    for row in rows:
        student_id = row["studentId"]
        quiz_id = row["quizId"]

        if student_id not in students_data:
            students_data[student_id] = {
                "Student Name": row.get("studentName", ""),
                "Roll Number": row.get("studentRollNo", ""),
            }
            # logger.debug(f"Student ID {student_id} added to students_data.")

        # Add quiz score info (only score, not total)
        students_data[student_id][quiz_id] = row["score"]

    # Convert to list for dataframe
    students_list = list(students_data.values())
    logger.debug(f"Converted students_data to list, {len(students_list)} students.")

    # Process quiz info to include formatted dates
    quiz_details = {}
    for quiz_id, info in quiz_info.items():
        title = info["title"]
        quiz_details[quiz_id] = {
            "title": title,
            "totalScore": info["totalScore"],
            "date": format_date(info["startTime"]) if info["startTime"] else "N/A",
        }
    logger.debug(f"Processed quiz details, {len(quiz_details)} quizzes.")

    return {
        "students": students_list,
        "quiz_details": quiz_details,
        "class_id": class_id,
        "student_count": len(students_list),
        "quiz_count": len(quiz_info),
        "start_date": start_date,
        "end_date": end_date,
    }


async def generate_excel_report(
    course_id: str,
    output_file: BytesIO = None,
    save_to_file: bool = True,
    start_date: str = None,
    end_date: str = None,
):
    """
    Generate an Excel report for the course.
    Returns a Dictionary with the Excel file as BytesIO Object and course code.

    Args:
        course_id: str - The course ID
        output_file: BytesIO - Optional output file to write to
        save_to_file: bool - Whether to save the report to a file
        start_date: str - Optional start date to filter quizzes (format: YYYY-MM-DD)
        end_date: str - Optional end date to filter quizzes (format: YYYY-MM-DD)
    """
    logger.debug(
        f"Generating Excel report for course ID: {course_id}, timeframe: {start_date} to {end_date}, save_to_file: {save_to_file}"
    )
    # Fetch the course report data with optional timeframe filtering
    report_data = await fetch_course_report(course_id, start_date, end_date)

    # Create Excel file in memory
    if not output_file:
        output = BytesIO()
        logger.debug("BytesIO object created for Excel file.")
    else:
        output = output_file

    # Create Excel writer and populate the sheet
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        course_code = populate_course_sheet(writer, course_id, report_data)

    # Reset file pointer to beginning
    output.seek(0)

    # Save the Excel file to a BytesIO object
    if save_to_file:
        await save_excel_report(output, course_code)
    return {"file": output, "course_code": course_code}


async def save_excel_report(
    excel_data: BytesIO, course_code: str, directory: str = "data/reports"
):
    """
    Save the Excel report (BytesIO object) to a file and return the file path.
    """
    logger.debug(
        f"Saving Excel report for course ID: {course_code} to directory: {directory}"
    )
    # Create directory if it doesn't exist
    os.makedirs(directory, exist_ok=True)

    # Create filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{directory}/course_{course_code}_report_{timestamp}.xlsx"

    # Save to file
    with open(filename, "wb") as file:
        file.write(excel_data.getvalue())

    logger.debug(f"Excel report saved to {filename}")
    return filename


if __name__ == "__main__":
    # Example usage
    import asyncio

    # Test fetch_course_report
    course_id = "155de92a02124783b8b4cca04de66819"
    # report = asyncio.run(fetch_course_report(course_id))
    # # print(report)
    # with open("data/course_report1.json", "w") as f:
    #     import json
    #     from app.utils.misc import DateTimeEncoder

    #     json.dump(report, f, indent=4, cls=DateTimeEncoder)

    # Test save_excel_report
    # directory = "data/reports"
    start_date = None
    end_date = "2025-03-16"
    # filename = asyncio.run(
    #     generate_excel_report(
    #         course_id, start_date=start_date, end_date=end_date, save_to_file=True
    #     )
    # )

    # Test generate_excel_class_report
    class_id = "cm48alxli00007ke7ch082mz0"
    report = asyncio.run(
        generate_excel_class_report(
            class_id, start_date=start_date, end_date=end_date, save_to_file=True
        )
    )
    # print(report)
